"""
Interface PyQt6 — Reiniciador de Workflow iiLex (Ramos Advogados).

Recursos:
    - Login com lembrar (DPAPI no Windows quando disponível)
    - Anexar planilha Excel + pré-visualização
    - Modo headless (opcional)
    - Iniciar / Pausar / Parar
    - Barra de progresso com contadores (✅ ⚠ ❌)
    - Tempo decorrido na status bar
    - Logs coloridos + salvos em arquivo
    - Relatório Excel automático ao final
    - Checkpoint: retomar execução interrompida
    - Notificação toast quando termina
    - Diálogo de configurações persistente
    - Validação opcional do formato CNJ

Como rodar:
    pip install -r requirements.txt
    python interface_iilex_qt.py
"""

from __future__ import annotations

import logging
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl
from PyQt6.QtCore import (
    QObject,
    QSize,
    QStringListModel,
    QThread,
    QTimer,
    Qt,
    QUrl,
    pyqtSignal,
    pyqtSlot,
)
from PyQt6.QtGui import (
    QColor,
    QDesktopServices,
    QFont,
    QIcon,
    QPixmap,
    QTextCharFormat,
    QTextCursor,
)
from PyQt6.QtWidgets import (
    QApplication,
    QCheckBox,
    QComboBox,
    QCompleter,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QFrame,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QSizePolicy,
    QSpinBox,
    QStatusBar,
    QStyle,
    QSystemTrayIcon,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

# Núcleo e utilitários do projeto
import checkpoint
import credentials
import report
from config import Settings, TIPOS_COMPROMISSO
from core_iilex import (
    CallbacksAutomacao,
    ConfigAutomacao,
    IilexAutomation,
    ResultadoProcesso,
    StatusProcesso,
    normalizar_texto,
    validar_cnj,
)


PASTA_BASE = Path(__file__).resolve().parent

# ============================================================
# VERSÃO / METADADOS
# ============================================================
VERSAO_APP = "1.2.0"
ANO_APP = "2026"


# ============================================================
# CAMINHO DE ASSETS (funciona tanto rodando .py quanto via PyInstaller)
# ============================================================
def caminho_asset(nome: str) -> str:
    """Resolve o caminho de um arquivo dentro de assets/.

    Quando rodando empacotado (PyInstaller --onefile), os arquivos
    bundleados ficam em sys._MEIPASS.
    """
    base = getattr(sys, "_MEIPASS", None)
    if base:
        candidato = Path(base) / "assets" / nome
        if candidato.exists():
            return str(candidato)
    return str(PASTA_BASE / "assets" / nome)


# ============================================================
# LEITURA DE EXCEL
# ============================================================
def detectar_coluna_processo(ws) -> tuple[int | None, str | None]:
    """Procura, no cabeçalho (linha 1), a coluna do 'Número do Processo'.

    Retorna (índice 1-based, texto do cabeçalho) ou (None, None) se não houver
    cabeçalho reconhecível. Serve pra auto-detectar a coluna certa em exports
    da Agenda do iiLex, onde o nº do processo NÃO fica na coluna A (fica, por
    ex., na coluna P = 'Número do Processo').
    """
    primeira = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not primeira:
        return None, None
    # Match forte (cabeçalho normalizado == um destes)
    exatos = {
        "numero do processo", "n do processo", "no do processo",
        "numero processo", "processo", "nº processo", "no processo",
    }
    fraco = None
    for i, cel in enumerate(primeira, start=1):
        if cel is None:
            continue
        h = normalizar_texto(str(cel))
        if h in exatos:
            return i, str(cel).strip()
        # Match fraco: contém "processo" mas não é "Tipo de Compromisso" etc.
        if fraco is None and "processo" in h and "tipo" not in h:
            fraco = (i, str(cel).strip())
    return fraco if fraco else (None, None)


def ler_processos_excel(
    caminho: str,
    coluna: int = 1,
    aba: str = "",
    ignorar_duplicados: bool = True,
) -> tuple[list[str], dict]:
    """Lê os números de processo de uma planilha.

    Estratégia:
      1. AUTO-DETECTA a coluna pelo cabeçalho 'Número do Processo' / 'Processo'
         (assim os exports da Agenda do iiLex funcionam direto, mesmo com o nº
         do processo lá na coluna P).
      2. Se não houver cabeçalho reconhecível, usa a `coluna` informada.

    Pula cabeçalhos e qualquer célula SEM dígitos (rótulos como 'Número do
    Processo'). Retorna (numeros, info), onde
    info = {'coluna': int, 'cabecalho': str | None, 'auto': bool}.
    """
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[aba] if aba and aba in wb.sheetnames else wb.active

    col_auto, cab_auto = detectar_coluna_processo(ws)
    col = col_auto if col_auto else coluna

    cabecalhos = {"processo", "nº processo", "numero", "número", "n°",
                  "numero do processo", "número do processo"}
    numeros: list[str] = []
    vistos: set[str] = set()
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < col or row[col - 1] is None:
            continue
        valor = str(row[col - 1]).strip()
        if not valor or valor.lower() in cabecalhos:
            continue
        # Nº de processo SEMPRE tem dígito → pula rótulos/cabeçalhos de texto.
        if not any(c.isdigit() for c in valor):
            continue
        if ignorar_duplicados and valor in vistos:
            continue
        vistos.add(valor)
        numeros.append(valor)
    wb.close()
    info = {"coluna": col, "cabecalho": cab_auto, "auto": bool(col_auto)}
    return numeros, info


# ============================================================
# WORKER (usa o core_iilex)
# ============================================================
@dataclass
class ParametrosTarefa:
    usuario: str
    senha: str
    processos: list[str]
    arquivo_excel: str
    settings: Settings
    retomar_de: int = 0


class Worker(QObject):
    """Adapter QObject sobre o IilexAutomation."""

    log_message = pyqtSignal(str, str)
    progress = pyqtSignal(int, int)
    status = pyqtSignal(str)
    resultado_individual = pyqtSignal(object)  # ResultadoProcesso
    finished = pyqtSignal(list)  # list[ResultadoProcesso]

    def __init__(self, params: ParametrosTarefa):
        super().__init__()
        self.params = params
        self._stop = False
        self._pause = False

    @pyqtSlot()
    def parar(self):
        self._stop = True

    @pyqtSlot(bool)
    def pausar(self, pause: bool):
        self._pause = pause

    @pyqtSlot()
    def executar(self):
        cfg = ConfigAutomacao(
            timeout=self.params.settings.timeout,
            headless=self.params.settings.headless,
            url_login=self.params.settings.url_login,
            url_contencioso=self.params.settings.url_contencioso,
            pausa_entre_processos=self.params.settings.pausa_entre_processos,
            tipo_compromisso_alvo=self.params.settings.tipo_compromisso_alvo,
            excluir_workflow=self.params.settings.excluir_workflow,
            relogin_minutos=self.params.settings.relogin_minutos,
        )
        callbacks = CallbacksAutomacao(
            on_log=lambda lvl, m: self.log_message.emit(lvl, m),
            on_progress=lambda i, t: self.progress.emit(i, t),
            on_status=lambda s: self.status.emit(s),
            on_resultado=self._on_resultado,
            is_stop_requested=lambda: self._stop,
            is_pause_requested=lambda: self._pause,
        )
        bot = IilexAutomation(cfg, callbacks)
        resultados = bot.executar(
            self.params.usuario,
            self.params.senha,
            self.params.processos,
            retomar_de=self.params.retomar_de,
        )
        self.finished.emit(resultados)

    def _on_resultado(self, r: ResultadoProcesso):
        self.resultado_individual.emit(r)


# ============================================================
# DIÁLOGO: PRÉ-VISUALIZAÇÃO DO EXCEL
# ============================================================
class DialogoPreview(QDialog):
    def __init__(self, numeros: list[str], arquivo: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Pré-visualização dos processos")
        self.resize(520, 480)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(f"<b>Arquivo:</b> {Path(arquivo).name}"))
        layout.addWidget(QLabel(f"<b>Total de processos:</b> {len(numeros)}"))

        invalidos = [n for n in numeros if not validar_cnj(n)]
        if invalidos:
            layout.addWidget(QLabel(
                f"<span style='color:#b45309'>⚠ {len(invalidos)} "
                f"número(s) fora do padrão CNJ (mas serão usados mesmo assim)</span>"
            ))

        tabela = QTableWidget(len(numeros), 2)
        tabela.setHorizontalHeaderLabels(["#", "Processo"])
        tabela.verticalHeader().setVisible(False)
        for i, n in enumerate(numeros):
            tabela.setItem(i, 0, QTableWidgetItem(str(i + 1)))
            item = QTableWidgetItem(n)
            if not validar_cnj(n):
                item.setForeground(QColor("#b45309"))
            tabela.setItem(i, 1, item)
        tabela.setColumnWidth(0, 50)
        tabela.horizontalHeader().setStretchLastSection(True)
        tabela.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(tabela)

        botoes = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        botoes.rejected.connect(self.reject)
        botoes.accepted.connect(self.accept)
        layout.addWidget(botoes)


# ============================================================
# DIÁLOGO: CONFIGURAÇÕES
# ============================================================
class DialogoConfiguracoes(QDialog):
    def __init__(self, settings: Settings, parent=None):
        super().__init__(parent)
        self.settings = settings
        self.setWindowTitle("Configurações")
        self.setMinimumWidth(500)

        layout = QVBoxLayout(self)
        form = QFormLayout()

        # Tema (logo no topo — primeira coisa que o usuário vê)
        self.cb_tema = QComboBox()
        self.cb_tema.addItem("Escuro (noturno)", "escuro")
        self.cb_tema.addItem("Claro", "claro")
        idx = 0 if settings.tema != "claro" else 1
        self.cb_tema.setCurrentIndex(idx)
        form.addRow("Tema:", self.cb_tema)

        self.sp_timeout = QSpinBox()
        self.sp_timeout.setRange(5, 120)
        self.sp_timeout.setValue(settings.timeout)
        self.sp_timeout.setSuffix(" s")
        form.addRow("Timeout do Selenium:", self.sp_timeout)

        self.sp_pausa = QSpinBox()
        self.sp_pausa.setRange(0, 30)
        self.sp_pausa.setValue(int(settings.pausa_entre_processos))
        self.sp_pausa.setSuffix(" s")
        form.addRow("Pausa entre processos:", self.sp_pausa)

        self.sp_coluna = QSpinBox()
        self.sp_coluna.setRange(1, 50)
        self.sp_coluna.setValue(settings.coluna_excel)
        form.addRow("Coluna do Excel (1 = A):", self.sp_coluna)

        self.ed_aba = QLineEdit(settings.aba_excel)
        self.ed_aba.setPlaceholderText("(vazio = aba ativa)")
        form.addRow("Aba do Excel:", self.ed_aba)

        self.chk_dup = QCheckBox("Ignorar processos duplicados")
        self.chk_dup.setChecked(settings.ignorar_duplicados)
        form.addRow("", self.chk_dup)

        self.chk_cnj = QCheckBox("Validar formato CNJ (avisar inválidos)")
        self.chk_cnj.setChecked(settings.validar_cnj)
        form.addRow("", self.chk_cnj)

        self.chk_pular_inv = QCheckBox("Pular processos com formato CNJ inválido")
        self.chk_pular_inv.setChecked(settings.pular_invalidos)
        form.addRow("", self.chk_pular_inv)

        self.chk_log = QCheckBox("Salvar logs em arquivo")
        self.chk_log.setChecked(settings.salvar_log_em_arquivo)
        form.addRow("", self.chk_log)

        self.chk_notify = QCheckBox("Notificar quando terminar")
        self.chk_notify.setChecked(settings.notificar_ao_terminar)
        form.addRow("", self.chk_notify)

        self.chk_retomar = QCheckBox("Perguntar para retomar checkpoint")
        self.chk_retomar.setChecked(settings.retomar_automatico)
        form.addRow("", self.chk_retomar)

        self.ed_url_login = QLineEdit(settings.url_login)
        form.addRow("URL de login:", self.ed_url_login)

        self.ed_url_cont = QLineEdit(settings.url_contencioso)
        form.addRow("URL do Contencioso:", self.ed_url_cont)

        layout.addLayout(form)

        info = QLabel(
            f"<i>Modo de credenciais: "
            f"{'🔒 DPAPI (Windows)' if credentials.dpapi_disponivel() else '⚠ base64 (instale pywin32 para DPAPI)'}"
            f"</i>"
        )
        info.setStyleSheet("color:#6b7280; padding-top:8px;")
        layout.addWidget(info)

        botoes = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save
            | QDialogButtonBox.StandardButton.Cancel
        )
        botoes.accepted.connect(self.accept)
        botoes.rejected.connect(self.reject)
        layout.addWidget(botoes)

    def settings_atualizadas(self) -> Settings:
        s = self.settings
        s.tema = self.cb_tema.currentData() or "escuro"
        s.timeout = self.sp_timeout.value()
        s.pausa_entre_processos = float(self.sp_pausa.value())
        s.coluna_excel = self.sp_coluna.value()
        s.aba_excel = self.ed_aba.text().strip()
        s.ignorar_duplicados = self.chk_dup.isChecked()
        s.validar_cnj = self.chk_cnj.isChecked()
        s.pular_invalidos = self.chk_pular_inv.isChecked()
        s.salvar_log_em_arquivo = self.chk_log.isChecked()
        s.notificar_ao_terminar = self.chk_notify.isChecked()
        s.retomar_automatico = self.chk_retomar.isChecked()
        s.url_login = self.ed_url_login.text().strip() or s.url_login
        s.url_contencioso = self.ed_url_cont.text().strip() or s.url_contencioso
        return s


# ============================================================
# QSS
# ============================================================
QSS_LIGHT = """
/* Paleta inspirada na marca Ramos Advogados:
   Navy:   #1F3559 (escuro do logo)
   Bege:   #B8A589 (claro do logo)
   Bege2:  #C9B79D (acento)
*/

QMainWindow, QWidget {
    background-color: #f5f7fa;
    color: #1f2937;
    font-family: "Segoe UI", Arial, sans-serif;
    font-size: 10pt;
}

/* ----------- HEADER ----------- */
QFrame#header {
    background-color: #1F3559;
    border: none;
    border-bottom: 3px solid #B8A589;
}
QFrame#header > QWidget,
QWidget#headerTitles,
QWidget#headerTitles > QWidget {
    background-color: transparent;
}
QFrame#header QLabel { color: #ffffff; background-color: transparent; }
QLabel#tituloHeader {
    color: #ffffff;
    font-size: 19pt;
    font-weight: 700;
    letter-spacing: 0.5px;
    background-color: transparent;
}
QLabel#subtituloHeader {
    color: #C9B79D;
    font-size: 9pt;
    font-weight: 500;
    letter-spacing: 0.3px;
    background-color: transparent;
}
QPushButton#btnHeaderConfig {
    background-color: rgba(255, 255, 255, 0.10);
    color: #ffffff;
    border: 1px solid rgba(255, 255, 255, 0.25);
    border-radius: 6px;
    padding: 8px 16px;
    font-weight: 600;
}
QPushButton#btnHeaderConfig:hover {
    background-color: rgba(255, 255, 255, 0.20);
    border: 1px solid rgba(255, 255, 255, 0.40);
}
QPushButton#btnHeaderConfig:pressed {
    background-color: rgba(255, 255, 255, 0.30);
}

/* ----------- FOOTER ----------- */
QFrame#footer {
    background-color: #ffffff;
    border: none;
    border-top: 2px solid #B8A589;
}
QFrame#footer > QWidget { background: transparent; }
QFrame#footer QLabel { background: transparent; }
QLabel#footerEmpresa {
    color: #1F3559;
    font-weight: 700;
    font-size: 10pt;
    background: transparent;
}
QLabel#footerCreditos {
    color: #94a3b8;
    font-size: 8pt;
    font-weight: 500;
    background: transparent;
}

/* ----------- GROUP BOX ----------- */
QGroupBox {
    background-color: #ffffff;
    border: 1px solid #e5e7eb;
    border-radius: 8px;
    margin-top: 14px;
    padding-top: 14px;
    font-weight: 600;
}
QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left;
    padding: 0 8px;
    color: #1F3559;
}

QLabel { color: #374151; }
QLabel#arquivoLabel { color: #6b7280; font-style: italic; }
QLabel#arquivoLabel[ativo="true"] {
    color: #111827; font-style: normal; font-weight: 600;
}
QLabel#contadores {
    font-family: "Cascadia Mono", "Consolas", monospace;
    font-size: 10pt;
    color: #1f2937;
    padding: 4px 0;
}

/* ----------- INPUTS ----------- */
QLineEdit, QSpinBox, QComboBox {
    background: #ffffff;
    border: 1px solid #d1d5db;
    border-radius: 6px;
    padding: 6px 10px;
    min-height: 22px;
}
QLineEdit:focus, QSpinBox:focus, QComboBox:focus {
    border: 1px solid #1F3559;
}

/* ----------- BOTÕES ----------- */
QPushButton {
    background-color: #ffffff;
    border: 1px solid #d1d5db;
    border-radius: 6px;
    padding: 6px 14px;
    min-height: 26px;
    color: #1f2937;
    font-weight: 600;
}
QPushButton:hover { background-color: #f3f4f6; }
QPushButton:pressed { background-color: #e5e7eb; }
QPushButton:disabled {
    background-color: #f3f4f6; color: #9ca3af; border: 1px solid #e5e7eb;
}

QPushButton#iniciar {
    background-color: #16a34a; color: white; border: 1px solid #15803d;
}
QPushButton#iniciar:hover { background-color: #15803d; }
QPushButton#iniciar:disabled {
    background-color: #d1fae5; color: #6ee7b7; border: 1px solid #a7f3d0;
}
QPushButton#parar {
    background-color: #dc2626; color: white; border: 1px solid #b91c1c;
}
QPushButton#parar:hover { background-color: #b91c1c; }
QPushButton#parar:disabled {
    background-color: #fecaca; color: #fca5a5; border: 1px solid #fecaca;
}
QPushButton#pausar {
    background-color: #d97706; color: white; border: 1px solid #b45309;
}
QPushButton#pausar:hover { background-color: #b45309; }
QPushButton#pausar:disabled {
    background-color: #fde68a; color: #fbbf24; border: 1px solid #fde68a;
}
QPushButton#pausar:checked { background-color: #f59e0b; }
QPushButton#anexar {
    background-color: #1F3559; color: white; border: 1px solid #142340;
}
QPushButton#anexar:hover { background-color: #142340; }
QPushButton#preview {
    background-color: #6b7280; color: white; border: 1px solid #4b5563;
}
QPushButton#preview:hover { background-color: #4b5563; }

QCheckBox { color: #374151; spacing: 6px; }

/* ----------- LOGS ----------- */
QTextEdit {
    background-color: #0f172a;
    color: #e2e8f0;
    border: 1px solid #1e293b;
    border-radius: 6px;
    font-family: "Cascadia Mono", "Consolas", monospace;
    font-size: 9pt;
    padding: 6px;
}

/* ----------- PROGRESS BAR (mais alta e bonita) ----------- */
QProgressBar {
    background-color: #e5e7eb;
    border: 1px solid #cbd5e1;
    border-radius: 8px;
    text-align: center;
    height: 26px;
    color: #ffffff;
    font-weight: 700;
    font-size: 10pt;
}
QProgressBar::chunk {
    border-radius: 7px;
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0    #1F3559,
        stop:0.6  #2563eb,
        stop:1    #16a34a);
}

/* ----------- STATUS BAR ----------- */
QStatusBar {
    background-color: #ffffff;
    border-top: 1px solid #e5e7eb;
    color: #4b5563;
}

/* ----------- TABELAS (preview) ----------- */
QTableWidget {
    background: #ffffff;
    border: 1px solid #e5e7eb;
    border-radius: 6px;
    gridline-color: #f3f4f6;
}
QHeaderView::section {
    background: #1F3559;
    color: white;
    padding: 6px;
    border: none;
    font-weight: 600;
}
"""


# ============================================================
# QSS — TEMA ESCURO (padrão)
# ============================================================
QSS_DARK = """
/* Paleta do tema escuro:
   Base:    #0f172a (slate-900) — fundo principal
   Card:    #1e293b (slate-800) — group boxes / inputs
   Border:  #334155 (slate-700)
   Text:    #e2e8f0 (slate-200)
   Muted:   #94a3b8 (slate-400)
   Navy:    #1F3559 (mantido — marca Ramos)
   Bege:    #B8A589 (acento)
*/

QMainWindow, QWidget {
    background-color: #0f172a;
    color: #e2e8f0;
    font-family: "Segoe UI", Arial, sans-serif;
    font-size: 10pt;
}

/* ----------- HEADER ----------- */
QFrame#header {
    background-color: #1F3559;
    border: none;
    border-bottom: 3px solid #B8A589;
}
QFrame#header > QWidget,
QWidget#headerTitles,
QWidget#headerTitles > QWidget {
    background-color: transparent;
}
QFrame#header QLabel { color: #ffffff; background-color: transparent; }
QLabel#tituloHeader {
    color: #ffffff;
    font-size: 19pt;
    font-weight: 700;
    letter-spacing: 0.5px;
    background-color: transparent;
}
QLabel#subtituloHeader {
    color: #C9B79D;
    font-size: 9pt;
    font-weight: 500;
    letter-spacing: 0.3px;
    background-color: transparent;
}
QPushButton#btnHeaderConfig {
    background-color: rgba(255, 255, 255, 0.10);
    color: #ffffff;
    border: 1px solid rgba(255, 255, 255, 0.25);
    border-radius: 6px;
    padding: 8px 16px;
    font-weight: 600;
}
QPushButton#btnHeaderConfig:hover {
    background-color: rgba(255, 255, 255, 0.20);
    border: 1px solid rgba(255, 255, 255, 0.40);
}
QPushButton#btnHeaderConfig:pressed {
    background-color: rgba(255, 255, 255, 0.30);
}

/* ----------- FOOTER ----------- */
QFrame#footer {
    background-color: #1e293b;
    border: none;
    border-top: 2px solid #B8A589;
}
QFrame#footer > QWidget { background: transparent; }
QFrame#footer QLabel { background: transparent; }
QLabel#footerEmpresa {
    color: #C9B79D;
    font-weight: 700;
    font-size: 10pt;
    background: transparent;
}
QLabel#footerCreditos {
    color: #64748b;
    font-size: 8pt;
    font-weight: 500;
    background: transparent;
}

/* ----------- GROUP BOX ----------- */
QGroupBox {
    background-color: #1e293b;
    border: 1px solid #334155;
    border-radius: 8px;
    margin-top: 14px;
    padding-top: 14px;
    font-weight: 600;
    color: #e2e8f0;
}
QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left;
    padding: 0 8px;
    color: #C9B79D;
    background-color: #1e293b;
}

QLabel { color: #cbd5e1; }
QLabel#arquivoLabel { color: #94a3b8; font-style: italic; }
QLabel#arquivoLabel[ativo="true"] {
    color: #f8fafc; font-style: normal; font-weight: 600;
}
QLabel#contadores {
    font-family: "Cascadia Mono", "Consolas", monospace;
    font-size: 10pt;
    color: #e2e8f0;
    padding: 4px 0;
}

/* ----------- INPUTS ----------- */
QLineEdit, QSpinBox, QComboBox {
    background: #0f172a;
    color: #e2e8f0;
    border: 1px solid #334155;
    border-radius: 6px;
    padding: 6px 10px;
    min-height: 22px;
    selection-background-color: #1F3559;
}
QLineEdit:focus, QSpinBox:focus, QComboBox:focus {
    border: 1px solid #B8A589;
}
QLineEdit:disabled, QSpinBox:disabled, QComboBox:disabled {
    background: #1e293b; color: #64748b;
}
QComboBox QAbstractItemView {
    background: #1e293b;
    color: #e2e8f0;
    border: 1px solid #334155;
    selection-background-color: #1F3559;
}

/* ----------- BOTÕES ----------- */
QPushButton {
    background-color: #1e293b;
    border: 1px solid #334155;
    border-radius: 6px;
    padding: 6px 14px;
    min-height: 26px;
    color: #e2e8f0;
    font-weight: 600;
}
QPushButton:hover { background-color: #334155; }
QPushButton:pressed { background-color: #475569; }
QPushButton:disabled {
    background-color: #0f172a; color: #475569; border: 1px solid #1e293b;
}

QPushButton#iniciar {
    background-color: #16a34a; color: white; border: 1px solid #15803d;
}
QPushButton#iniciar:hover { background-color: #15803d; }
QPushButton#iniciar:disabled {
    background-color: #064e3b; color: #6ee7b7; border: 1px solid #065f46;
}
QPushButton#parar {
    background-color: #dc2626; color: white; border: 1px solid #b91c1c;
}
QPushButton#parar:hover { background-color: #b91c1c; }
QPushButton#parar:disabled {
    background-color: #7f1d1d; color: #fca5a5; border: 1px solid #991b1b;
}
QPushButton#pausar {
    background-color: #d97706; color: white; border: 1px solid #b45309;
}
QPushButton#pausar:hover { background-color: #b45309; }
QPushButton#pausar:disabled {
    background-color: #78350f; color: #fbbf24; border: 1px solid #92400e;
}
QPushButton#pausar:checked { background-color: #f59e0b; }
QPushButton#anexar {
    background-color: #1F3559; color: white; border: 1px solid #B8A589;
}
QPushButton#anexar:hover { background-color: #2a4670; }
QPushButton#preview {
    background-color: #475569; color: white; border: 1px solid #64748b;
}
QPushButton#preview:hover { background-color: #64748b; }

QCheckBox { color: #cbd5e1; spacing: 6px; }
QCheckBox::indicator {
    width: 16px; height: 16px;
    border: 1px solid #475569;
    border-radius: 3px;
    background: #0f172a;
}
QCheckBox::indicator:checked {
    background: #B8A589;
    border: 1px solid #B8A589;
}

/* ----------- LOGS (mantém terminal style) ----------- */
QTextEdit {
    background-color: #020617;
    color: #e2e8f0;
    border: 1px solid #1e293b;
    border-radius: 6px;
    font-family: "Cascadia Mono", "Consolas", monospace;
    font-size: 9pt;
    padding: 6px;
}

/* ----------- PROGRESS BAR ----------- */
QProgressBar {
    background-color: #1e293b;
    border: 1px solid #334155;
    border-radius: 8px;
    text-align: center;
    height: 26px;
    color: #ffffff;
    font-weight: 700;
    font-size: 10pt;
}
QProgressBar::chunk {
    border-radius: 7px;
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0    #1F3559,
        stop:0.6  #2563eb,
        stop:1    #16a34a);
}

/* ----------- STATUS BAR ----------- */
QStatusBar {
    background-color: #1e293b;
    border-top: 1px solid #334155;
    color: #94a3b8;
}
QStatusBar QLabel { color: #94a3b8; }

/* ----------- DIÁLOGOS ----------- */
QDialog {
    background-color: #0f172a;
    color: #e2e8f0;
}

/* ----------- TABELAS (preview) ----------- */
QTableWidget {
    background: #1e293b;
    color: #e2e8f0;
    border: 1px solid #334155;
    border-radius: 6px;
    gridline-color: #334155;
    selection-background-color: #1F3559;
}
QHeaderView::section {
    background: #1F3559;
    color: white;
    padding: 6px;
    border: none;
    font-weight: 600;
}

/* ----------- SCROLLBAR (sutil, combina com dark) ----------- */
QScrollBar:vertical {
    background: #1e293b;
    width: 10px;
    border-radius: 5px;
}
QScrollBar::handle:vertical {
    background: #475569;
    border-radius: 5px;
    min-height: 24px;
}
QScrollBar::handle:vertical:hover { background: #64748b; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
"""


# ============================================================
# Selector de tema
# ============================================================
def montar_qss(tema: str) -> str:
    """Retorna o QSS correto pro tema escolhido ('escuro' | 'claro')."""
    return QSS_LIGHT if tema == "claro" else QSS_DARK


# ============================================================
# JANELA PRINCIPAL
# ============================================================
class JanelaPrincipal(QMainWindow):
    CORES_LOG = {
        "INFO": QColor("#93c5fd"),
        "WARNING": QColor("#fcd34d"),
        "ERROR": QColor("#fca5a5"),
        "OK": QColor("#86efac"),
    }

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Reiniciador de Workflow iiLex — Ramos Advogados")
        self.resize(780, 700)
        self.setMinimumSize(680, 560)

        # Estado
        self.settings = Settings.carregar()
        self.arquivo_excel: str | None = None
        self.processos_carregados: list[str] = []
        self.thread: QThread | None = None
        self.worker: Worker | None = None
        self.inicio_execucao: float = 0.0
        self.contadores = {s: 0 for s in StatusProcesso}
        self.resultados_atuais: list[ResultadoProcesso] = []
        self.indice_corrente = 0  # último índice processado (para checkpoint)
        self.tray: QSystemTrayIcon | None = None
        self._file_log_handler: logging.FileHandler | None = None
        self._path_relatorio: str | None = None   # relatório Excel completo (tudo)
        self._path_sucesso: str | None = None     # planilha só dos concluídos
        self._path_pendencias: str | None = None  # planilha dos que precisam revisar

        # UI
        self._criar_ui()
        self._aplicar_tema()
        self._criar_tray()
        self._carregar_credenciais()
        self._setup_logging_arquivo()

        # Timer pro tempo decorrido
        self.timer_elapsed = QTimer(self)
        self.timer_elapsed.setInterval(1000)
        self.timer_elapsed.timeout.connect(self._tick_elapsed)

    # ---------------------------------------------------------
    # UI
    # ---------------------------------------------------------
    def _criar_ui(self):
        # Define o ícone da janela (taskbar + barra de título)
        icon_path = caminho_asset("ramos.ico")
        if Path(icon_path).exists():
            self.setWindowIcon(QIcon(icon_path))

        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ====================================================
        # HEADER estilizado (navy + logo + título + config)
        # ====================================================
        root.addWidget(self._build_header())

        # ====================================================
        # ÁREA DE CONTEÚDO (com margens normais)
        # ====================================================
        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(18, 16, 18, 12)
        layout.setSpacing(10)
        root.addWidget(content, 1)

        # Credenciais
        grp_cred = QGroupBox("Credenciais")
        v_cred = QVBoxLayout(grp_cred)
        linha_user = QHBoxLayout()
        lbl_user = QLabel("Login:"); lbl_user.setFixedWidth(60)
        self.input_user = QLineEdit()
        self.input_user.setPlaceholderText("seu.usuario")
        linha_user.addWidget(lbl_user); linha_user.addWidget(self.input_user)
        v_cred.addLayout(linha_user)
        linha_pass = QHBoxLayout()
        lbl_pass = QLabel("Senha:"); lbl_pass.setFixedWidth(60)
        self.input_pass = QLineEdit()
        self.input_pass.setEchoMode(QLineEdit.EchoMode.Password)
        self.input_pass.setPlaceholderText("••••••••")
        linha_pass.addWidget(lbl_pass); linha_pass.addWidget(self.input_pass)
        v_cred.addLayout(linha_pass)
        self.chk_lembrar = QCheckBox("Lembrar login e senha (DPAPI quando disponível)")
        v_cred.addWidget(self.chk_lembrar)
        layout.addWidget(grp_cred)

        # Planilha + headless
        grp_arq = QGroupBox("Planilha")
        v_arq = QVBoxLayout(grp_arq)
        linha_arq = QHBoxLayout()
        self.lbl_arquivo = QLabel("Nenhum arquivo selecionado")
        self.lbl_arquivo.setObjectName("arquivoLabel")
        self.lbl_arquivo.setProperty("ativo", False)
        self.btn_anexar = QPushButton("Anexar Excel")
        self.btn_anexar.setObjectName("anexar")
        self.btn_anexar.clicked.connect(self._anexar_excel)
        self.btn_preview = QPushButton("Pré-visualizar")
        self.btn_preview.setObjectName("preview")
        self.btn_preview.setEnabled(False)
        self.btn_preview.clicked.connect(self._mostrar_preview)
        linha_arq.addWidget(self.lbl_arquivo, 1)
        linha_arq.addWidget(self.btn_preview)
        linha_arq.addWidget(self.btn_anexar)
        v_arq.addLayout(linha_arq)
        self.chk_headless = QCheckBox("Rodar em segundo plano (sem janela do Chrome)")
        self.chk_headless.setChecked(self.settings.headless)
        v_arq.addWidget(self.chk_headless)
        self.chk_excluir_wf = QCheckBox(
            "Excluir o WorkFlow ao entrar (somente compromissos em negrito e não concluídos)")
        self.chk_excluir_wf.setChecked(self.settings.excluir_workflow)
        v_arq.addWidget(self.chk_excluir_wf)
        layout.addWidget(grp_arq)

        # Tipo de compromisso-alvo (combo pesquisável) — na tela principal
        # porque o usuário troca o tipo com frequência.
        layout.addWidget(self._criar_grupo_compromisso())

        # Ações
        h_btn = QHBoxLayout()
        self.btn_iniciar = QPushButton("▶ Iniciar")
        self.btn_iniciar.setObjectName("iniciar")
        self.btn_iniciar.clicked.connect(self._iniciar)
        self.btn_pausar = QPushButton("⏸ Pausar")
        self.btn_pausar.setObjectName("pausar")
        self.btn_pausar.setCheckable(True)
        self.btn_pausar.setEnabled(False)
        self.btn_pausar.clicked.connect(self._toggle_pausar)
        self.btn_parar = QPushButton("■ Parar")
        self.btn_parar.setObjectName("parar")
        self.btn_parar.setEnabled(False)
        self.btn_parar.clicked.connect(self._parar)
        self.btn_relatorio = QPushButton("📊 Abrir relatório")
        self.btn_relatorio.setObjectName("preview")
        self.btn_relatorio.setToolTip("Abre o relatório Excel completo da execução")
        self.btn_relatorio.setVisible(False)
        self.btn_relatorio.clicked.connect(self._abrir_relatorio)
        self.btn_sucesso = QPushButton("✅ Abrir concluídos")
        self.btn_sucesso.setObjectName("preview")
        self.btn_sucesso.setToolTip(
            "Planilha só dos concluídos com sucesso (entrou / WorkFlow reiniciado)")
        self.btn_sucesso.setVisible(False)
        self.btn_sucesso.clicked.connect(self._abrir_sucesso)
        self.btn_pendencias = QPushButton("📋 Abrir revisar")
        self.btn_pendencias.setObjectName("preview")
        self.btn_pendencias.setToolTip(
            "Planilha dos que precisam revisar: sem compromisso, já concluído, "
            "múltiplos, não encontrado, erro ou pulado")
        self.btn_pendencias.setVisible(False)
        self.btn_pendencias.clicked.connect(self._abrir_pendencias)
        self.btn_limpar = QPushButton("🗑 Limpar logs")
        self.btn_limpar.clicked.connect(lambda: self.txt_log.clear())
        h_btn.addWidget(self.btn_iniciar)
        h_btn.addWidget(self.btn_pausar)
        h_btn.addWidget(self.btn_parar)
        h_btn.addStretch()
        h_btn.addWidget(self.btn_relatorio)
        h_btn.addWidget(self.btn_sucesso)
        h_btn.addWidget(self.btn_pendencias)
        h_btn.addWidget(self.btn_limpar)
        layout.addLayout(h_btn)

        # Progresso + contadores
        self.progresso = QProgressBar()
        self.progresso.setRange(0, 1); self.progresso.setValue(0)
        self.progresso.setFormat("Pronto")
        layout.addWidget(self.progresso)

        self.lbl_contadores = QLabel(self._formatar_contadores())
        self.lbl_contadores.setObjectName("contadores")
        layout.addWidget(self.lbl_contadores)

        # Logs
        grp_log = QGroupBox("Logs")
        v_log = QVBoxLayout(grp_log)
        self.txt_log = QTextEdit(); self.txt_log.setReadOnly(True)
        v_log.addWidget(self.txt_log)
        layout.addWidget(grp_log, 1)

        # ====================================================
        # FOOTER estilizado
        # ====================================================
        root.addWidget(self._build_footer())

        # Status bar (do QMainWindow, fica logo abaixo do footer)
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.lbl_elapsed = QLabel("⏱ 00:00")
        self.status_bar.addPermanentWidget(self.lbl_elapsed)
        self.status_bar.showMessage("Pronto.")

    # ---------------------------------------------------------
    # TIPO DE COMPROMISSO (combo pesquisável da tela principal)
    # ---------------------------------------------------------
    def _criar_grupo_compromisso(self) -> QGroupBox:
        """Grupo da tela principal com o seletor de Tipo de Compromisso.

        Combo EDITÁVEL + pesquisável (ignora acento e maiúscula/minúscula).
        O valor escolhido é o compromisso-alvo que a automação procura na
        Agenda. Ficava nas Configurações; foi movido pra cá porque o usuário
        troca o tipo com frequência.
        """
        grp = QGroupBox("Tipo de Compromisso")
        v = QVBoxLayout(grp)

        self.cb_tipo_comp = QComboBox()
        self.cb_tipo_comp.setEditable(True)
        self.cb_tipo_comp.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.cb_tipo_comp.addItems(TIPOS_COMPROMISSO)

        # Busca "contém" ignorando ACENTO e caixa (o QCompleter puro não faz).
        # Por isso: UnfilteredPopupCompletion + filtro à mão num QStringListModel.
        self._tipos_model = QStringListModel(list(TIPOS_COMPROMISSO), self)
        completer = QCompleter(self._tipos_model, self.cb_tipo_comp)
        completer.setCompletionMode(
            QCompleter.CompletionMode.UnfilteredPopupCompletion
        )
        completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        self.cb_tipo_comp.setCompleter(completer)
        self._tipos_completer = completer
        self.cb_tipo_comp.lineEdit().textEdited.connect(self._filtrar_tipos)

        # Pré-seleciona o valor salvo (casando por texto normalizado).
        self._selecionar_tipo(self.settings.tipo_compromisso_alvo)

        self.cb_tipo_comp.lineEdit().setPlaceholderText(
            "Digite para buscar…  ex.: Subsídios")
        self.cb_tipo_comp.setToolTip(
            "A automação ENTRA no compromisso cujo 'Tipo de Compromisso' na "
            "Agenda for EXATAMENTE este (ignora acento/maiúscula).\n"
            "Escolha da lista ou digite para buscar/usar um tipo novo.\n"
            "Variações com sufixo NÃO entram — ex.: 'SOLICITAÇÃO DE SUBSÍDIOS - CEF'."
        )
        v.addWidget(self.cb_tipo_comp)
        return grp

    def _selecionar_tipo(self, valor: str):
        """Pré-seleciona `valor` no combo casando por texto normalizado
        (acento/caixa). Assim 'Solicitação de Subsídios' acha
        'SOLICITAÇÃO DE SUBSIDIOS'. Se não existir, usa como texto livre."""
        idx = self.cb_tipo_comp.findText(valor, Qt.MatchFlag.MatchFixedString)
        if idx < 0:
            alvo = normalizar_texto(valor)
            for i, t in enumerate(TIPOS_COMPROMISSO):
                if normalizar_texto(t) == alvo:
                    idx = i
                    break
        if idx >= 0:
            self.cb_tipo_comp.setCurrentIndex(idx)
        else:
            self.cb_tipo_comp.setCurrentText(valor)

    def _filtrar_tipos(self, texto: str):
        """Filtra o popup do combo ignorando ACENTO e maiúscula/minúscula.
        Texto vazio = lista completa de volta."""
        alvo = normalizar_texto(texto)
        if alvo:
            filtrados = [t for t in TIPOS_COMPROMISSO if alvo in normalizar_texto(t)]
        else:
            filtrados = list(TIPOS_COMPROMISSO)
        self._tipos_model.setStringList(filtrados)
        self._tipos_completer.complete()

    # ---------------------------------------------------------
    # HEADER
    # ---------------------------------------------------------
    def _build_header(self) -> QFrame:
        header = QFrame()
        header.setObjectName("header")
        header.setFixedHeight(86)
        header.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed
        )
        h = QHBoxLayout(header)
        h.setContentsMargins(24, 14, 24, 14)
        h.setSpacing(18)

        # Logo (R com barras) — tamanho compacto pro header
        logo_label = QLabel()
        logo_label.setObjectName("headerLogo")
        logo_pix = QPixmap(caminho_asset("R_logo.png"))
        if not logo_pix.isNull():
            logo_pix = logo_pix.scaledToHeight(
                50, Qt.TransformationMode.SmoothTransformation
            )
            logo_label.setPixmap(logo_pix)
        # Largura proporcional à altura escalada — evita o logo "comer" o título
        logo_label.setFixedHeight(58)
        logo_label.setAlignment(
            Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter
        )
        logo_label.setSizePolicy(
            QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed
        )
        h.addWidget(logo_label, alignment=Qt.AlignmentFlag.AlignVCenter)

        # Bloco do título — empilhado verticalmente, centralizado.
        # Define objectName + attribute pra garantir que o fundo
        # do container fique transparente (e o navy do header apareça).
        titulos_widget = QWidget()
        titulos_widget.setObjectName("headerTitles")
        titulos_widget.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        titulos = QVBoxLayout(titulos_widget)
        titulos.setContentsMargins(0, 0, 0, 0)
        titulos.setSpacing(2)
        titulos.addStretch()
        titulo = QLabel("Reiniciador de Workflow iiLex")
        titulo.setObjectName("tituloHeader")
        subtitulo = QLabel("Ramos Advogados")
        subtitulo.setObjectName("subtituloHeader")
        titulos.addWidget(titulo)
        titulos.addWidget(subtitulo)
        titulos.addStretch()
        h.addWidget(titulos_widget, 1, alignment=Qt.AlignmentFlag.AlignVCenter)

        # Botão configurações (estilo translúcido sobre o navy)
        self.btn_config = QPushButton("⚙  Configurações")
        self.btn_config.setObjectName("btnHeaderConfig")
        self.btn_config.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_config.clicked.connect(self._abrir_configuracoes)
        h.addWidget(self.btn_config, alignment=Qt.AlignmentFlag.AlignVCenter)

        return header

    # ---------------------------------------------------------
    # FOOTER
    # ---------------------------------------------------------
    def _build_footer(self) -> QFrame:
        footer = QFrame()
        footer.setObjectName("footer")
        footer.setFixedHeight(52)
        footer.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed
        )
        h = QHBoxLayout(footer)
        h.setContentsMargins(20, 8, 20, 8)
        h.setSpacing(12)

        # Logo R pequena
        logo_label = QLabel()
        logo_label.setObjectName("footerLogo")
        logo_pix = QPixmap(caminho_asset("R_logo.png"))
        if not logo_pix.isNull():
            logo_pix = logo_pix.scaledToHeight(
                32, Qt.TransformationMode.SmoothTransformation
            )
            logo_label.setPixmap(logo_pix)
        logo_label.setFixedHeight(36)
        logo_label.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        h.addWidget(logo_label, alignment=Qt.AlignmentFlag.AlignVCenter)

        # Texto da empresa + tagline em duas linhas
        empresa_box = QVBoxLayout()
        empresa_box.setContentsMargins(0, 0, 0, 0)
        empresa_box.setSpacing(0)
        empresa = QLabel("Ramos Advogados")
        empresa.setObjectName("footerEmpresa")
        tagline = QLabel("Tecnologia jurídica")
        tagline.setObjectName("footerCreditos")
        empresa_box.addWidget(empresa)
        empresa_box.addWidget(tagline)
        h.addLayout(empresa_box)

        h.addStretch()

        # Créditos / versão à direita
        versao_box = QVBoxLayout()
        versao_box.setContentsMargins(0, 0, 0, 0)
        versao_box.setSpacing(0)
        versao_box.setAlignment(Qt.AlignmentFlag.AlignRight)
        nome_app = QLabel("Reiniciador de Workflow iiLex")
        nome_app.setObjectName("footerEmpresa")
        nome_app.setAlignment(Qt.AlignmentFlag.AlignRight)
        versao_txt = QLabel(f"v{VERSAO_APP}  •  © {ANO_APP}")
        versao_txt.setObjectName("footerCreditos")
        versao_txt.setAlignment(Qt.AlignmentFlag.AlignRight)
        versao_box.addWidget(nome_app)
        versao_box.addWidget(versao_txt)
        h.addLayout(versao_box)

        return footer

    def _criar_tray(self):
        if not QSystemTrayIcon.isSystemTrayAvailable():
            return
        icon_path = caminho_asset("ramos.ico")
        if Path(icon_path).exists():
            icon = QIcon(icon_path)
        else:
            icon = self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon)
        self.tray = QSystemTrayIcon(icon, self)
        self.tray.setToolTip("Reiniciador de Workflow iiLex")
        self.tray.show()

    # ---------------------------------------------------------
    # Anexar / Pré-visualizar Excel
    # ---------------------------------------------------------
    def _anexar_excel(self):
        caminho, _ = QFileDialog.getOpenFileName(
            self,
            "Selecione a planilha de processos",
            "",
            "Planilhas Excel (*.xlsx *.xlsm *.xls);;Todos os arquivos (*.*)",
        )
        if not caminho:
            return
        try:
            numeros, info = ler_processos_excel(
                caminho,
                coluna=self.settings.coluna_excel,
                aba=self.settings.aba_excel,
                ignorar_duplicados=self.settings.ignorar_duplicados,
            )
        except Exception as e:
            QMessageBox.critical(self, "Erro ao ler Excel", str(e))
            return

        self.arquivo_excel = caminho
        self.processos_carregados = numeros
        self.lbl_arquivo.setText(f"📄 {Path(caminho).name} — {len(numeros)} processos")
        self.lbl_arquivo.setProperty("ativo", True)
        self.lbl_arquivo.style().unpolish(self.lbl_arquivo)
        self.lbl_arquivo.style().polish(self.lbl_arquivo)
        self.btn_preview.setEnabled(True)

        # Transparência: qual coluna foi usada (auto-detecção pelo cabeçalho).
        from openpyxl.utils import get_column_letter
        letra = get_column_letter(info["coluna"])
        if info["auto"]:
            self._log("INFO",
                f'Coluna {letra} detectada pelo cabeçalho "{info["cabecalho"]}" '
                f"— {len(numeros)} processo(s) carregado(s).")
        else:
            self._log("INFO",
                f"Lendo coluna {letra} — {len(numeros)} processo(s) carregado(s).")
        if not numeros:
            self._log("WARNING",
                "Nenhum número de processo encontrado. Confira se a planilha "
                "tem uma coluna 'Número do Processo'.")

        if self.settings.validar_cnj:
            invalidos = [n for n in numeros if not validar_cnj(n)]
            if invalidos:
                self._log("WARNING",
                    f"{len(invalidos)} processo(s) fora do padrão CNJ"
                    + (" (serão pulados — vide Configurações)" if self.settings.pular_invalidos else "")
                )

    def _mostrar_preview(self):
        if not self.processos_carregados or not self.arquivo_excel:
            return
        DialogoPreview(self.processos_carregados, self.arquivo_excel, self).exec()

    def _abrir_relatorio(self):
        """Abre o relatório Excel completo da última execução."""
        if self._path_relatorio and Path(self._path_relatorio).exists():
            QDesktopServices.openUrl(QUrl.fromLocalFile(self._path_relatorio))
        else:
            QMessageBox.information(
                self, "Sem relatório",
                "Nenhum relatório disponível ainda.")

    def _abrir_sucesso(self):
        """Abre a planilha só dos concluídos com sucesso."""
        if self._path_sucesso and Path(self._path_sucesso).exists():
            QDesktopServices.openUrl(QUrl.fromLocalFile(self._path_sucesso))
        else:
            QMessageBox.information(
                self, "Sem concluídos",
                "Nenhuma planilha de concluídos disponível.")

    def _abrir_pendencias(self):
        """Abre a planilha dos que precisam revisar."""
        if self._path_pendencias and Path(self._path_pendencias).exists():
            QDesktopServices.openUrl(QUrl.fromLocalFile(self._path_pendencias))
        else:
            QMessageBox.information(
                self, "Sem pendências",
                "Nenhuma planilha de revisão disponível.")

    # ---------------------------------------------------------
    # Configurações
    # ---------------------------------------------------------
    def _abrir_configuracoes(self):
        dlg = DialogoConfiguracoes(self.settings, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            tema_antigo = self.settings.tema
            self.settings = dlg.settings_atualizadas()
            self.settings.salvar()
            self._setup_logging_arquivo()
            # Reaplica o tema se mudou
            if self.settings.tema != tema_antigo:
                self._aplicar_tema()
            self._log("INFO", "Configurações salvas.")

    def _aplicar_tema(self):
        """Aplica o QSS do tema atual em toda a aplicação."""
        app = QApplication.instance()
        if app is not None:
            app.setStyleSheet(montar_qss(self.settings.tema))

    # ---------------------------------------------------------
    # Iniciar / Parar / Pausar
    # ---------------------------------------------------------
    def _iniciar(self):
        usuario = self.input_user.text().strip()
        senha = self.input_pass.text().strip()
        if not usuario or not senha:
            QMessageBox.warning(self, "Campos obrigatórios",
                "Preencha o login e a senha antes de iniciar.")
            return
        if not self.processos_carregados:
            QMessageBox.warning(self, "Sem processos",
                "Anexe uma planilha Excel com pelo menos um processo.")
            return
        if self.thread is not None and self.thread.isRunning():
            QMessageBox.information(self, "Já em execução",
                "A automação já está rodando.")
            return

        # Persistência de credenciais
        if self.chk_lembrar.isChecked():
            credentials.salvar(credentials.Credenciais(usuario=usuario, senha=senha))
        else:
            credentials.apagar()

        # Aplica opções escolhidas na UI principal
        self.settings.headless = self.chk_headless.isChecked()
        self.settings.excluir_workflow = self.chk_excluir_wf.isChecked()
        self.settings.tipo_compromisso_alvo = (
            self.cb_tipo_comp.currentText().strip()
            or self.settings.tipo_compromisso_alvo
        )
        self.settings.salvar()  # persiste o tipo escolhido p/ a próxima vez

        # Filtra inválidos se configurado
        processos = self.processos_carregados
        if self.settings.pular_invalidos:
            originais = len(processos)
            processos = [p for p in processos if validar_cnj(p)]
            if len(processos) < originais:
                self._log("WARNING",
                    f"{originais - len(processos)} processo(s) inválido(s) pulado(s).")
        if not processos:
            QMessageBox.warning(self, "Sem processos válidos",
                "Nenhum processo válido para executar.")
            return

        # Checkpoint — pergunta se quer retomar
        retomar_de = 0
        if self.settings.retomar_automatico:
            cp = checkpoint.carregar()
            if checkpoint.pode_retomar(cp, processos, self.arquivo_excel or ""):
                resp = QMessageBox.question(
                    self,
                    "Retomar checkpoint",
                    f"Encontrei um checkpoint de uma execução anterior:\n\n"
                    f"• Arquivo: {Path(cp.arquivo_excel).name}\n"
                    f"• Já processados: {cp.proximo_indice} de {cp.total}\n"
                    f"• Salvo em: {cp.salvo_em}\n\n"
                    f"Deseja retomar de onde parou?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                )
                if resp == QMessageBox.StandardButton.Yes:
                    retomar_de = cp.proximo_indice
                else:
                    checkpoint.apagar()

        # Reset de estado
        self.contadores = {s: 0 for s in StatusProcesso}
        self.resultados_atuais = []
        self.indice_corrente = retomar_de
        self._path_relatorio = None
        self.btn_relatorio.setVisible(False)
        self._path_sucesso = None
        self.btn_sucesso.setVisible(False)
        self._path_pendencias = None
        self.btn_pendencias.setVisible(False)
        self.lbl_contadores.setText(self._formatar_contadores())
        self.progresso.setRange(0, len(processos))
        self.progresso.setValue(retomar_de)
        self.progresso.setFormat(f"{retomar_de}/{len(processos)} processos")

        # Worker + Thread
        params = ParametrosTarefa(
            usuario=usuario,
            senha=senha,
            processos=processos,
            arquivo_excel=self.arquivo_excel or "",
            settings=self.settings,
            retomar_de=retomar_de,
        )
        self.thread = QThread()
        self.worker = Worker(params)
        self.worker.moveToThread(self.thread)

        self.thread.started.connect(self.worker.executar)
        self.worker.log_message.connect(self._append_log)
        self.worker.progress.connect(self._atualizar_progresso)
        self.worker.status.connect(self.status_bar.showMessage)
        self.worker.resultado_individual.connect(self._on_resultado)
        self.worker.finished.connect(self._on_finished)
        self.worker.finished.connect(self.thread.quit)
        self.worker.finished.connect(self.worker.deleteLater)
        self.thread.finished.connect(self.thread.deleteLater)

        # UI estado
        self.btn_iniciar.setEnabled(False)
        self.btn_parar.setEnabled(True)
        self.btn_pausar.setEnabled(True)
        self.btn_anexar.setEnabled(False)
        self.btn_config.setEnabled(False)
        self.chk_headless.setEnabled(False)
        self.chk_excluir_wf.setEnabled(False)
        self.progresso.setFormat("Iniciando...")

        # Tempo decorrido
        self.inicio_execucao = time.time()
        self.lbl_elapsed.setText("⏱ 00:00")
        self.timer_elapsed.start()

        self.thread.start()

    def _toggle_pausar(self, checked: bool):
        if not self.worker:
            return
        self.worker.pausar(checked)
        self.btn_pausar.setText("▶ Retomar" if checked else "⏸ Pausar")

    def _parar(self):
        if self.worker:
            self.worker.parar()
            self._log("WARNING", "Sinal de parada enviado...")
            self.status_bar.showMessage("Parando...")

    @pyqtSlot(list)
    def _on_finished(self, resultados: list):
        self.timer_elapsed.stop()
        self.btn_iniciar.setEnabled(True)
        self.btn_parar.setEnabled(False)
        self.btn_pausar.setEnabled(False)
        self.btn_pausar.setChecked(False)
        self.btn_pausar.setText("⏸ Pausar")
        self.btn_anexar.setEnabled(True)
        self.btn_config.setEnabled(True)
        self.chk_headless.setEnabled(True)
        self.chk_excluir_wf.setEnabled(True)

        # Gera relatório se houver pelo menos um resultado
        path_relat = None
        if resultados:
            pasta = self.settings.pasta_relatorios_resolvida(PASTA_BASE / "relatorios")
            try:
                path_relat = report.gerar_relatorio(resultados, pasta)
                self._path_relatorio = str(path_relat)
                self.btn_relatorio.setVisible(True)
                self.btn_relatorio.setEnabled(True)
                self._log("OK", f"Relatório gerado: {path_relat}")
            except Exception as e:
                self._log("ERROR", f"Falha ao gerar relatório: {e}")
            # Planilha só dos CONCLUÍDOS com sucesso (entrou / WF reiniciado)
            try:
                path_ok = report.gerar_planilha_sucesso(resultados, pasta)
                if path_ok:
                    self._path_sucesso = str(path_ok)
                    self.btn_sucesso.setVisible(True)
                    self.btn_sucesso.setEnabled(True)
                    self._log("OK", f"Planilha de concluídos gerada: {path_ok}")
            except Exception as e:
                self._log("ERROR", f"Falha ao gerar planilha de concluídos: {e}")
            # Planilha separada dos que precisam REVISAR (sem compromisso, já
            # concluído, múltiplos, não encontrado, erro, pulado)
            try:
                path_pend = report.gerar_planilha_pendencias(resultados, pasta)
                if path_pend:
                    self._path_pendencias = str(path_pend)
                    self.btn_pendencias.setVisible(True)
                    self.btn_pendencias.setEnabled(True)
                    self._log("WARNING",
                              f"Planilha de revisão gerada: {path_pend}")
            except Exception as e:
                self._log("ERROR", f"Falha ao gerar planilha de revisão: {e}")

        # Checkpoint: se concluiu tudo, apaga; senão, salva onde parou
        total = len(self.processos_carregados)
        if self.indice_corrente >= total:
            checkpoint.apagar()
        else:
            cp = checkpoint.Checkpoint(
                arquivo_excel=self.arquivo_excel or "",
                hash_processos=checkpoint.calcular_hash(self.processos_carregados),
                total=total,
                proximo_indice=self.indice_corrente,
            )
            checkpoint.salvar(cp)
            self._log("WARNING",
                f"Checkpoint salvo no índice {self.indice_corrente}/{total}.")

        # Notificação
        if self.settings.notificar_ao_terminar and self.tray:
            resumo = self._resumo_contagem()
            self.tray.showMessage(
                "Reiniciador de Workflow iiLex concluído",
                f"{resumo}" + (f"\nRelatório: {Path(path_relat).name}" if path_relat else ""),
                QSystemTrayIcon.MessageIcon.Information,
                5000,
            )

        self.worker = None

    # ---------------------------------------------------------
    # Recebimento de eventos do worker
    # ---------------------------------------------------------
    @pyqtSlot(int, int)
    def _atualizar_progresso(self, atual: int, total: int):
        self.indice_corrente = atual
        if total <= 0:
            self.progresso.setRange(0, 1); self.progresso.setValue(0)
            self.progresso.setFormat("Sem processos"); return
        self.progresso.setRange(0, total)
        self.progresso.setValue(atual)
        self.progresso.setFormat(f"{atual}/{total} processos")

        # Salva checkpoint a cada processo
        if atual > 0 and self.arquivo_excel:
            cp = checkpoint.Checkpoint(
                arquivo_excel=self.arquivo_excel,
                hash_processos=checkpoint.calcular_hash(self.processos_carregados),
                total=len(self.processos_carregados),
                proximo_indice=atual,
            )
            checkpoint.salvar(cp)

    @pyqtSlot(object)
    def _on_resultado(self, resultado: ResultadoProcesso):
        self.resultados_atuais.append(resultado)
        self.contadores[resultado.status] = self.contadores.get(resultado.status, 0) + 1
        self.lbl_contadores.setText(self._formatar_contadores())

    @pyqtSlot(str, str)
    def _append_log(self, level: str, message: str):
        self._log(level, message)

    # ---------------------------------------------------------
    # Logging
    # ---------------------------------------------------------
    def _log(self, level: str, message: str):
        cor = self.CORES_LOG.get(level, QColor("#e2e8f0"))
        timestamp = time.strftime("%H:%M:%S")

        cursor = self.txt_log.textCursor()
        cursor.movePosition(QTextCursor.MoveOperation.End)

        fmt = QTextCharFormat()
        fmt.setForeground(QColor("#64748b"))
        cursor.insertText(f"{timestamp} ", fmt)

        fmt = QTextCharFormat()
        fmt.setForeground(cor)
        fmt.setFontWeight(QFont.Weight.Bold)
        cursor.insertText(f"{level:<7} ", fmt)

        fmt = QTextCharFormat()
        fmt.setForeground(QColor("#e2e8f0"))
        cursor.insertText(f"{message}\n", fmt)

        self.txt_log.setTextCursor(cursor)
        self.txt_log.ensureCursorVisible()

        # Replica no arquivo
        if self._file_log_handler:
            logging.getLogger("iilex").log(
                {"INFO": logging.INFO, "WARNING": logging.WARNING,
                 "ERROR": logging.ERROR, "OK": logging.INFO}.get(level, logging.INFO),
                message,
            )

    def _setup_logging_arquivo(self):
        # Remove handler antigo se existir
        logger = logging.getLogger("iilex")
        if self._file_log_handler:
            logger.removeHandler(self._file_log_handler)
            self._file_log_handler.close()
            self._file_log_handler = None

        if not self.settings.salvar_log_em_arquivo:
            return

        pasta = self.settings.pasta_logs_resolvida(PASTA_BASE / "logs")
        pasta.mkdir(parents=True, exist_ok=True)
        nome = f"iilex_{datetime.now():%Y-%m-%d}.log"
        handler = logging.FileHandler(pasta / nome, encoding="utf-8")
        handler.setFormatter(
            logging.Formatter("%(asctime)s | %(levelname)-7s | %(message)s", "%H:%M:%S")
        )
        logger.setLevel(logging.INFO)
        logger.addHandler(handler)
        self._file_log_handler = handler

    # ---------------------------------------------------------
    # Tempo decorrido / contadores
    # ---------------------------------------------------------
    def _tick_elapsed(self):
        delta = int(time.time() - self.inicio_execucao)
        m, s = divmod(delta, 60)
        h, m = divmod(m, 60)
        if h:
            self.lbl_elapsed.setText(f"⏱ {h:02d}:{m:02d}:{s:02d}")
        else:
            self.lbl_elapsed.setText(f"⏱ {m:02d}:{s:02d}")

    def _formatar_contadores(self) -> str:
        c = self.contadores
        return (
            f"🔄 {c.get(StatusProcesso.WORKFLOW_REINICIADO, 0)} workflow reiniciado  |  "
            f"⚠ {c.get(StatusProcesso.SEM_COMPROMISSO, 0)} sem compromisso  |  "
            f"🔁 {c.get(StatusProcesso.JA_CONCLUIDO, 0)} já concluído  |  "
            f"🔎 {c.get(StatusProcesso.NAO_ENCONTRADO, 0)} não encontrado  |  "
            f"❌ {c.get(StatusProcesso.ERRO, 0)} erros  |  "
            f"⏭ {c.get(StatusProcesso.PULADO, 0)} pulados"
        )

    def _resumo_contagem(self) -> str:
        c = self.contadores
        return (
            f"{c.get(StatusProcesso.WORKFLOW_REINICIADO, 0)} workflow reiniciado, "
            f"{c.get(StatusProcesso.ENTROU, 0)} entrou, "
            f"{c.get(StatusProcesso.SEM_COMPROMISSO, 0)} sem compromisso, "
            f"{c.get(StatusProcesso.JA_CONCLUIDO, 0)} já concluídos, "
            f"{c.get(StatusProcesso.MULTIPLOS, 0)} múltiplos, "
            f"{c.get(StatusProcesso.NAO_ENCONTRADO, 0)} não encontrados, "
            f"{c.get(StatusProcesso.ERRO, 0)} erros"
        )

    # ---------------------------------------------------------
    # Credenciais
    # ---------------------------------------------------------
    def _carregar_credenciais(self):
        c = credentials.carregar()
        if c:
            self.input_user.setText(c.usuario)
            self.input_pass.setText(c.senha)
            self.chk_lembrar.setChecked(True)

    # ---------------------------------------------------------
    # Ciclo de vida
    # ---------------------------------------------------------
    def closeEvent(self, event):
        if self.thread is not None and self.thread.isRunning():
            r = QMessageBox.question(
                self,
                "Automação em execução",
                "A automação ainda está rodando. Deseja parar e fechar?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if r != QMessageBox.StandardButton.Yes:
                event.ignore(); return
            if self.worker:
                self.worker.parar()
            self.thread.quit()
            self.thread.wait(3000)
        event.accept()


# ============================================================
# ENTRADA
# ============================================================
def main():
    app = QApplication(sys.argv)
    app.setFont(QFont("Segoe UI", 10))
    app.setApplicationName("Reiniciador de Workflow iiLex")
    app.setOrganizationName("Ramos Advogados")

    # Ícone da aplicação (aparece também na taskbar do Windows)
    icon_path = caminho_asset("ramos.ico")
    if Path(icon_path).exists():
        app.setWindowIcon(QIcon(icon_path))

    # Aplica o tema persistido (default = escuro)
    settings_iniciais = Settings.carregar()
    app.setStyleSheet(montar_qss(settings_iniciais.tema))

    janela = JanelaPrincipal()
    janela.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
