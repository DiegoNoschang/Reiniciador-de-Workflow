"""
Geração de relatório Excel ao final da automação.

Layout:
    Aba 'Resultados':
        Processo | Status | Mensagem | Hora
    Aba 'Resumo':
        Total / Entrou / Sem compromisso / Já concluído / Múltiplos / Erros / Pulados

Cores por status:
    Entrou          → verde claro
    Sem compromisso → amarelo claro
    Já concluído    → azul claro
    Múltiplos       → laranja claro
    Erro            → vermelho claro
    Pulado          → cinza
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core_iilex import ResultadoProcesso, StatusProcesso


# Cores (preenchimento das células)
CORES = {
    StatusProcesso.WORKFLOW_REINICIADO: PatternFill("solid", fgColor="92D050"),  # verde forte
    StatusProcesso.ENTROU: PatternFill("solid", fgColor="C6EFCE"),         # verde
    StatusProcesso.SEM_COMPROMISSO: PatternFill("solid", fgColor="FFEB9C"),  # amarelo
    StatusProcesso.JA_CONCLUIDO: PatternFill("solid", fgColor="BDD7EE"),    # azul
    StatusProcesso.MULTIPLOS: PatternFill("solid", fgColor="FCE4D6"),       # laranja
    StatusProcesso.NAO_ENCONTRADO: PatternFill("solid", fgColor="E1D5E7"),  # lilás
    StatusProcesso.ERRO: PatternFill("solid", fgColor="FFC7CE"),           # vermelho
    StatusProcesso.PULADO: PatternFill("solid", fgColor="E7E6E6"),         # cinza
}

CABECALHO_FILL = PatternFill("solid", fgColor="1F4E78")
CABECALHO_FONT = Font(bold=True, color="FFFFFF")


def gerar_relatorio(
    resultados: Iterable[ResultadoProcesso],
    destino: Path,
) -> Path:
    """Gera um .xlsx no caminho `destino` com os resultados.

    Se o caminho for um diretório, salva como
    'iilex_relatorio_YYYY-MM-DD_HHMMSS.xlsx' dentro dele.
    Retorna o caminho final do arquivo.
    """
    resultados = list(resultados)
    destino = Path(destino)

    if destino.is_dir() or destino.suffix == "":
        destino.mkdir(parents=True, exist_ok=True)
        nome = f"iilex_relatorio_{datetime.now():%Y-%m-%d_%H%M%S}.xlsx"
        destino = destino / nome

    wb = Workbook()
    _aba_resultados(wb, resultados)
    _aba_resumo(wb, resultados)
    wb.save(destino)
    return destino


# Status "concluídos com sucesso" (a automação entrou / reiniciou o WorkFlow)
STATUS_SUCESSO = (
    StatusProcesso.WORKFLOW_REINICIADO,
    StatusProcesso.ENTROU,
)

# Status que precisam de atenção (revisar / reprocessar) = tudo que NÃO foi
# concluído com sucesso: sem compromisso, já concluído, múltiplos, não
# encontrado, erro e pulado.
STATUS_PENDENCIA = (
    StatusProcesso.SEM_COMPROMISSO,
    StatusProcesso.JA_CONCLUIDO,
    StatusProcesso.MULTIPLOS,
    StatusProcesso.NAO_ENCONTRADO,
    StatusProcesso.ERRO,
    StatusProcesso.PULADO,
)


def _gerar_planilha_filtrada(
    resultados: Iterable[ResultadoProcesso],
    destino: Path,
    status_alvo: tuple,
    titulo_aba: str,
    prefixo_arquivo: str,
) -> Path | None:
    """Gera um .xlsx só com os resultados cujo status está em `status_alvo`.

    Colunas: Processo | Status | Mensagem (a 1ª coluna é o número, então a
    planilha pode ser reanexada direto numa nova rodada). Retorna o caminho
    do arquivo, ou None se não houver nenhum resultado nesses status.
    """
    sel = [r for r in resultados if r.status in status_alvo]
    if not sel:
        return None

    destino = Path(destino)
    if destino.is_dir() or destino.suffix == "":
        destino.mkdir(parents=True, exist_ok=True)
        nome = f"{prefixo_arquivo}_{datetime.now():%Y-%m-%d_%H%M%S}.xlsx"
        destino = destino / nome

    wb = Workbook()
    ws = wb.active
    ws.title = titulo_aba

    headers = ["Processo", "Status", "Mensagem"]
    for col, h in enumerate(headers, 1):
        cel = ws.cell(row=1, column=col, value=h)
        cel.fill = CABECALHO_FILL
        cel.font = CABECALHO_FONT
        cel.alignment = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(sel, start=2):
        ws.cell(row=i, column=1, value=r.numero)
        ws.cell(row=i, column=2, value=r.status.value)
        ws.cell(row=i, column=3, value=r.mensagem)
        fill = CORES.get(r.status)
        if fill:
            for col in range(1, 4):
                ws.cell(row=i, column=col).fill = fill

    for col, w in enumerate([35, 24, 60], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    wb.save(destino)
    return destino


def gerar_planilha_pendencias(
    resultados: Iterable[ResultadoProcesso],
    destino: Path,
) -> Path | None:
    """Planilha SÓ com os que precisam de atenção (revisar/reprocessar):
    sem compromisso, já concluído, múltiplos, não encontrado, erro, pulado."""
    return _gerar_planilha_filtrada(
        resultados, destino, STATUS_PENDENCIA, "Revisar", "iilex_revisar")


def gerar_planilha_sucesso(
    resultados: Iterable[ResultadoProcesso],
    destino: Path,
) -> Path | None:
    """Planilha SÓ com os concluídos com sucesso (entrou / WorkFlow reiniciado)."""
    return _gerar_planilha_filtrada(
        resultados, destino, STATUS_SUCESSO, "Concluídos", "iilex_concluidos")


def _aba_resultados(wb: Workbook, resultados: list[ResultadoProcesso]):
    ws = wb.active
    ws.title = "Resultados"

    headers = ["#", "Processo", "Status", "Mensagem", "Hora"]
    for col, h in enumerate(headers, 1):
        celula = ws.cell(row=1, column=col, value=h)
        celula.fill = CABECALHO_FILL
        celula.font = CABECALHO_FONT
        celula.alignment = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(resultados, 1):
        linha = i + 1
        ws.cell(row=linha, column=1, value=i)
        ws.cell(row=linha, column=2, value=r.numero)
        cel_status = ws.cell(row=linha, column=3, value=r.status.value)
        ws.cell(row=linha, column=4, value=r.mensagem)
        ws.cell(row=linha, column=5, value=r.timestamp)

        fill = CORES.get(r.status)
        if fill:
            for col in range(1, 6):
                ws.cell(row=linha, column=col).fill = fill

    # Auto-tamanho aproximado de colunas
    larguras = [5, 35, 15, 60, 12]
    for col, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"


def _aba_resumo(wb: Workbook, resultados: list[ResultadoProcesso]):
    ws = wb.create_sheet("Resumo")

    contagem = {s: 0 for s in StatusProcesso}
    for r in resultados:
        contagem[r.status] += 1
    total = len(resultados)

    linhas = [
        ("Total de processos", total),
        ("WorkFlow reiniciado", contagem[StatusProcesso.WORKFLOW_REINICIADO]),
        ("Entrou (sem excluir)", contagem[StatusProcesso.ENTROU]),
        ("Sem compromisso do tipo", contagem[StatusProcesso.SEM_COMPROMISSO]),
        ("Já concluído", contagem[StatusProcesso.JA_CONCLUIDO]),
        ("Múltiplos pendentes", contagem[StatusProcesso.MULTIPLOS]),
        ("Processo não encontrado", contagem[StatusProcesso.NAO_ENCONTRADO]),
        ("Erros", contagem[StatusProcesso.ERRO]),
        ("Pulados", contagem[StatusProcesso.PULADO]),
        ("", ""),
        ("Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
    ]

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22

    titulo = ws.cell(row=1, column=1, value="Resumo da automação iiLex")
    titulo.font = Font(bold=True, size=14, color="1F4E78")
    ws.cell(row=1, column=2, value="")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    for i, (label, valor) in enumerate(linhas, start=3):
        cel_label = ws.cell(row=i, column=1, value=label)
        cel_val = ws.cell(row=i, column=2, value=valor)
        if label and not label.startswith("Gerado"):
            cel_label.font = Font(bold=True)
        cel_val.alignment = Alignment(horizontal="right" if isinstance(valor, int) else "left")

    # Pinta os números do resumo com as mesmas cores
    # (linhas começam em 3: 3=Total, 4=WorkFlow reiniciado, 5=Entrou,
    #  6=Sem comp., 7=Já concl., 8=Múltiplos, 9=Não encontrado,
    #  10=Erros, 11=Pulados)
    cores_resumo = {
        4: CORES[StatusProcesso.WORKFLOW_REINICIADO],
        5: CORES[StatusProcesso.ENTROU],
        6: CORES[StatusProcesso.SEM_COMPROMISSO],
        7: CORES[StatusProcesso.JA_CONCLUIDO],
        8: CORES[StatusProcesso.MULTIPLOS],
        9: CORES[StatusProcesso.NAO_ENCONTRADO],
        10: CORES[StatusProcesso.ERRO],
        11: CORES[StatusProcesso.PULADO],
    }
    for linha, fill in cores_resumo.items():
        ws.cell(row=linha, column=1).fill = fill
        ws.cell(row=linha, column=2).fill = fill
